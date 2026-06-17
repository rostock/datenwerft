import csv
from codecs import BOM_UTF8, iterdecode
from decimal import Decimal

from django.db import connections, transaction
from django.forms import Textarea
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from psycopg2.extras import execute_values
from psycopg2.sql import SQL, Identifier

from ..apps import StadtbereichskatalogConfig as appConfig
from ..utils import is_stadtbereichskatalog_user


def add_app_context_elements(context):
  """
  adds global app related elements to a context and returns it

  :param context: context
  :return: context with global app related elements added
  """
  context['app'] = {
    'name': getattr(appConfig, 'name', appConfig.name),
    'verbose_name': getattr(appConfig, 'verbose_name', appConfig.name),
    'description': getattr(appConfig, 'description', ''),
  }
  return context


def add_model_context_elements(context, model):
  """
  adds model related elements to a context and returns it

  :param context: context
  :param model: model
  :return: context with model related elements added
  """
  context['model_verbose_name'] = model._meta.verbose_name
  context['model_verbose_name_plural'] = model._meta.verbose_name_plural
  return context


def add_permissions_context_elements(context, user):
  """
  adds permissions related elements to a context and returns it

  :param context: context
  :param user: user
  :return: context with permissions related elements added
  """
  permissions = {
    'is_stadtbereichskatalog_user': is_stadtbereichskatalog_user(user),
  }
  if user.is_superuser:
    permissions = {key: True for key in permissions}
  context.update(permissions)
  return context


def assign_widget(field):
  """
  creates corresponding form field (widget) to passed model field and returns it

  :param field: form field
  :return: corresponding form field (widget) to passed model field
  """
  form_field = field.formfield()
  # handle inputs
  if hasattr(form_field.widget, 'input_type'):
    if form_field.widget.input_type == 'checkbox':
      form_field.widget.attrs['class'] = 'form-check-input'
    elif form_field.widget.input_type == 'select':
      form_field.widget.attrs['class'] = 'form-select'
    else:
      form_field.widget.attrs['class'] = 'form-control'
    if form_field.widget.input_type == 'number':
      form_field.widget.attrs['min'] = 0
  # handle text areas
  elif issubclass(form_field.widget.__class__, Textarea):
    form_field.widget.attrs['class'] = 'form-control'
    form_field.widget.attrs['rows'] = 10
  return form_field


def clean_headers(headers):
  """
  cleans passed headers

  :param headers: headers
  :return: passed headers, cleaned
  """
  cleaned_headers = []
  for header in headers:
    if header:
      cleaned_headers.append(header.replace('\ufeff', '').strip())
  return cleaned_headers


def parse_decimal(value):
  """
  parses passed value into a decimal value

  :param value: value to be parsed
  :return: passed value, parsed into a decimal value
  """
  value = value.replace(',', '.')
  return Decimal(value)


# type conversion registry
TYPE_CONVERTERS = {
  'integer': int,
  'numeric': parse_decimal,
}


def convert_value(value, pg_type):
  """
  converts passed value to passed PostgreSQL data type according to type conversion registry

  :param value: value to be converted
  :param pg_type: PostgreSQL data type
  :return: passed value, converted according to type conversion registry
  """
  if value in (None, ''):
    return None
  converter = TYPE_CONVERTERS.get(pg_type)
  if not converter:
    return value
  try:
    return converter(value)
  except Exception:
    # last-resort: only try string conversion if needed
    return converter(str(value))


def data_to_csv(
  schema_name, table_name, year_filter=None, area_filter=None, election_filter=None, standard=True
):
  """
  creates and returns CSV with data from a query
  based on passed table within passed database schema
  (and, optionally, passed filters)

  :param schema_name: name of database schema
  :param table_name: name of database table
  :param year_filter: optional year filter
  :param area_filter: optional area filter
  :param election_filter: optional election filter
  :param standard: standard CSV	(True) or Excel friendly CSV (False)?
  :return: CSV with data from a query based on passed table within passed database schema
  (and, optionally, passed filters)
  """

  # get data
  columns, rows = get_export_data(
    schema_name, table_name, year_filter, area_filter, election_filter
  )

  # prepare HTTP response with CSV headers
  response = HttpResponse(
    content_type='text/csv; charset=utf-8',
  )

  # set CSV name and prepare as file attachment
  filename = f'{schema_name}__{table_name}'
  if year_filter:
    filename += f'__{year_filter}'
  if area_filter:
    filename += f'__{area_filter}'
  if election_filter:
    filename += f'__{election_filter}'
  filename += '.csv'
  response['Content-Disposition'] = f'attachment; filename="{filename}"'

  if standard:
    # prepare standard CSV
    # important: lineterminator='\n' forces LF
    csv_writer = csv.writer(
      response,
      delimiter=',',
      quotechar='"',
      quoting=csv.QUOTE_MINIMAL,
      lineterminator='\n',
    )
  else:
    # prepare Excel friendly CSV
    # important: UTF-8 BOM
    response.write('\ufeff')
    # important: lineterminator='\r\n' forces CRLF
    csv_writer = csv.writer(
      response,
      delimiter=';',
      quotechar='"',
      quoting=csv.QUOTE_MINIMAL,
      lineterminator='\r\n',
    )

  # no data? return empty CSV
  if not rows:
    return response

  # add header row to CSV
  csv_writer.writerow(columns)

  # add data rows to CSV
  if standard:
    csv_writer.writerows(rows)
  else:
    # Excel-ify values for Excel friendly CSV
    csv_writer.writerows(
      [[excel_value(value) for value in row] for row in rows]
    )

  return response


def data_to_excel(
  schema_name, table_name, year_filter=None, area_filter=None, election_filter=None
):
  """
  creates and returns XSLX with data from a query
  based on passed table within passed database schema
  (and, optionally, passed filters)

  :param schema_name: name of database schema
  :param table_name: name of database table
  :param year_filter: optional year filter
  :param area_filter: optional area filter
  :param election_filter: optional election filter
  :return: XSLX with data from a query based on passed table within passed database schema
  (and, optionally, passed filters)
  """

  # get data
  columns, rows = get_export_data(
    schema_name, table_name, year_filter, area_filter, election_filter
  )

  # prepare HTTP response with XLSX headers
  response = HttpResponse(
    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
  )

  # set XLSX name and prepare as file attachment
  filename = f'{schema_name}__{table_name}.xlsx'
  response['Content-Disposition'] = f'attachment; filename="{filename}"'

  # create workbook
  wb = Workbook()
  ws = wb.active
  ws.title = 'Export'

  # no data? return empty XSLX
  if not rows:
    wb.save(response)
    return response

  # add header row to XSLX
  ws.append(columns)

  # add data rows to XSLX
  for row in rows:
    ws.append(row)

  # save workbook to response
  wb.save(response)

  return response


def delete_data(schema_name, table_name, year_filter=None, election_filter=None):
  """
  deletes data from passed table within passed database schema according to passed filters

  :param schema_name: name of database schema
  :param table_name: name of database table
  :param year_filter: optional year filter
  :param election_filter: optional election filter
  :return: results of deletion of data
  from passed table within passed database schema according to passed filters
  """

  # dynamic where clause
  conditions, params = [], []
  if year_filter:
    conditions.append('jahr = %s')
    params.append(int(year_filter))
  if election_filter:
    election = election_filter.split('__')
    election_election, election_year = election[0], election[1]
    conditions.append('wahlart_id = %s')
    params.append(election_election)
    conditions.append('wahljahr = %s')
    params.append(int(election_year))
  where_sql = ''
  if conditions:
    where_sql = ' AND '.join(conditions)

  # build query
  query = SQL("""
      DELETE FROM {}.{}
  """).format(Identifier(schema_name), Identifier(table_name))
  if where_sql:
    query += SQL(' WHERE ') + SQL(where_sql)

  # execute query
  try:
    connection = connections['stadtbereichskatalog']
    with connection.cursor() as cursor:
      cursor.execute(query, params)
    return {'success': True, 'deleted_rows': cursor.rowcount}
  except Exception as e:
    return {'success': False, 'errors': [{'row': 0, 'message': str(e)}]}


def detect_dialect(file_obj):
  """
  detects and returns dialect of passed file

  :param file_obj: file
  :return: dialect of passed file
  """
  sample = file_obj.read(4096)
  file_obj.seek(0)

  # handle UTF-8 BOM
  if sample.startswith(BOM_UTF8):
    sample = sample[len(BOM_UTF8) :]
  try:
    dialect = csv.Sniffer().sniff(sample.decode('utf-8'), delimiters=';,')
  except csv.Error:
    # fallback to strict default
    dialect = csv.excel
    dialect.delimiter = ','

  return dialect


def detect_file_type(file_obj):
  """
  detects and returns type of passed file

  :param file_obj: file
  :return type of passed file
  """
  name = file_obj.name.lower()
  if name.endswith('.xlsx'):
    return 'xlsx'
  if name.endswith('.csv'):
    return 'csv'
  return 'unknown'


def excel_value(value):
  """
  Excel-ifies passed value and returns it

  :param value: original value
  :return Excel-ified version of passed value
  """
  if isinstance(value, Decimal):
    return str(value).replace('.', ',')
  return value


def get_csv_reader(file_obj):
  """
  gets unified CSV reader factory for passed file

  :param file_obj: file
  :return: unified CSV reader factory for passed file as well as dialect of passed file
  """
  dialect = detect_dialect(file_obj)
  decoded = iterdecode(file_obj, 'utf-8-sig')
  reader = csv.DictReader(decoded, dialect=dialect)
  return reader, dialect


def get_database_columns(schema_name, table_name):
  """
  gets all columns of passed table within passed database schema and returns them

  :param schema_name: name of database schema
  :param table_name: name of database table
  :return: all columns of passed table within passed database schema
  """
  connection = connections['stadtbereichskatalog']
  with connection.cursor() as cursor:
    # columns
    cursor.execute(
      """
      SELECT
       c.column_name,
       c.data_type,
       c.is_nullable
        FROM information_schema.columns c
         WHERE c.table_schema = %s
         AND c.table_name = %s
          ORDER BY c.ordinal_position
    """,
      [schema_name, table_name],
    )
    columns = cursor.fetchall()

    # primary keys
    cursor.execute(
      """
      SELECT
       kcu.column_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu
         ON tc.constraint_name = kcu.constraint_name
         AND tc.table_schema = kcu.table_schema
          WHERE tc.constraint_type = 'PRIMARY KEY'
          AND tc.table_schema = %s
          AND tc.table_name = %s
    """,
      [schema_name, table_name],
    )
    primary_keys = {row[0] for row in cursor.fetchall()}

    result = []
    for name, dtype, nullable in columns:
      is_nullable = nullable == 'YES'
      result.append(
        {
          'name': name,
          'type': dtype,
          'required': not is_nullable,
          'primary_key': name in primary_keys,
        }
      )
    return result


def get_database_data(schema_name, table_name, order_by=None):
  """
  gets all data of passed table within passed database schema and returns them

  :param schema_name: name of database schema
  :param table_name: name of database table
  :param order_by: name(s) of column(s) to order by
  :return: all data of passed table within passed database schema
  """

  # build query
  query = SQL("""
      SELECT *
      FROM {}.{}
  """).format(Identifier(schema_name), Identifier(table_name))
  if order_by:
    query += SQL(' ORDER BY ') + SQL(order_by)

  # execute query
  connection = connections['stadtbereichskatalog']
  with connection.cursor() as cursor:
    cursor.execute(query)
    columns = [col[0] for col in cursor.description]
    rows = []
    for record in cursor.fetchall():
      rows.append(dict(zip(columns, record)))
    return rows


def get_database_schemas():
  """
  gets all database schemas and returns their names

  :return: names of all database schemas
  """
  connection = connections['stadtbereichskatalog']
  with connection.cursor() as cursor:
    cursor.execute("""
      SELECT schema_name
       FROM information_schema.schemata
        WHERE schema_name NOT IN (
         'geo', 'grafana', 'information_schema', 'metadaten', 'pg_catalog', 'pg_toast', 'public'
        )
         ORDER BY schema_name
    """)
    return [row[0] for row in cursor.fetchall()]


def get_database_tables(schema_name):
  """
  gets all tables of passed database schema and returns their names

  :param schema_name: name of database schema
  :return: names of all tables of passed database schema
  """
  connection = connections['stadtbereichskatalog']
  with connection.cursor() as cursor:
    cursor.execute(
      """
      SELECT table_name
       FROM information_schema.tables
        WHERE table_schema = %s
        AND table_type = 'BASE TABLE'
         ORDER BY table_name
    """,
      [schema_name],
    )
    return [row[0] for row in cursor.fetchall()]


def get_distinct_areas(schema_name, table_name):
  """
  gets all distinct areas of passed table within passed database schema and returns them

  :param schema_name: name of database schema
  :param table_name: name of database table
  :return: all distinct areas of passed table within passed database schema
  """
  connection = connections['stadtbereichskatalog']
  with connection.cursor() as cursor:
    cursor.execute(
      SQL(
        """
      SELECT s.kuerzel, s.name
       FROM {}.{} d
       JOIN public.stadtbereiche s ON d.stadtbereich = s.kuerzel
        GROUP BY s.kuerzel, s.name
         ORDER BY s.sortierung
    """
      ).format(Identifier(schema_name), Identifier(table_name))
    )
    areas = cursor.fetchall()
    result = []
    for code, name in areas:
      result.append(
        {
          'code': code,
          'name': name,
        }
      )
    return result


def get_distinct_elections(schema_name, table_name):
  """
  gets all distinct elections of passed table within passed database schema and returns them

  :param schema_name: name of database schema
  :param table_name: name of database table
  :return: all distinct elections of passed table within passed database schema
  """
  connection = connections['stadtbereichskatalog']
  with connection.cursor() as cursor:
    cursor.execute(
      SQL(
        """
      SELECT w.id AS wahl_id, w.name, d.wahljahr
       FROM {}.{} d
       JOIN public.wahlarten w ON d.wahlart_id = w.id
        GROUP BY w.id, w.name, d.wahljahr
         ORDER BY w.name, d.wahljahr DESC
    """
      ).format(Identifier(schema_name), Identifier(table_name))
    )
    elections = cursor.fetchall()
    result = []
    for wahl_id, name, wahljahr in elections:
      result.append(
        {
          'election': wahl_id,
          'name': name,
          'year': wahljahr,
        }
      )
    return result


def get_distinct_years(schema_name, table_name):
  """
  gets all distinct years of passed table within passed database schema and returns them

  :param schema_name: name of database schema
  :param table_name: name of database table
  :return: all distinct years of passed table within passed database schema
  """
  connection = connections['stadtbereichskatalog']
  column = 'wahljahr' if schema_name == 'wahlen' else 'jahr'
  with connection.cursor() as cursor:
    cursor.execute(  # correct
      SQL(
        """
      SELECT DISTINCT {}
       FROM {}.{}
        ORDER BY {} DESC
    """
      ).format(
        Identifier(column), Identifier(schema_name), Identifier(table_name), Identifier(column)
      )
    )
    return [row[0] for row in cursor.fetchall()]


def get_export_data(
  schema_name, table_name, year_filter=None, area_filter=None, election_filter=None
):
  """
  gets data of passed table within passed database schema according to passed filters
  and returns them

  :param schema_name: name of database schema
  :param table_name: name of database table
  :param year_filter: optional year filter
  :param area_filter: optional area filter
  :param election_filter: optional election filter
  :return: data of passed table within passed database schema
  """

  # dynamic where clause
  conditions, params = [], []
  if year_filter:
    conditions.append('jahr = %s')
    params.append(int(year_filter))
  if area_filter:
    conditions.append('stadtbereich = %s')
    params.append(area_filter)
  if election_filter:
    election = election_filter.split('__')
    election_election, election_year = election[0], election[1]
    conditions.append('wahlart_id = %s')
    params.append(election_election)
    conditions.append('wahljahr = %s')
    params.append(int(election_year))
  where_sql = ''
  if conditions:
    where_sql = ' AND '.join(conditions)

  # build query
  query = SQL("""
      SELECT *
      FROM {}.{}
  """).format(Identifier(schema_name), Identifier(table_name))
  if where_sql:
    query += SQL(' WHERE ') + SQL(where_sql)

  # execute query
  connection = connections['stadtbereichskatalog']
  with connection.cursor() as cursor:
    cursor.execute(query, params)
    rows = cursor.fetchall()
    columns = [col[0] for col in cursor.description]
    return columns, rows


def get_model_objects(model, count=False):
  """
  either gets all objects of passed model and returns them
  or counts objects of passed model and returns the count

  :param model: model
  :param count: return objects count instead of objects?
  :return: either all objects of passed model or objects count of passed model
  """
  objects = model.objects.all()
  return objects.count() if count else objects


def import_data(schema_name, table_name, columns, rows):
  """
  imports passed rows into passed columns of passed table within passed database schema

  :param schema_name: name of database schema
  :param table_name: name of database table
  :param columns: insert columns
  :param rows: rows to insert
  :return: results of import of passed rows
  into passed columns of passed table within passed database schema
  """

  quoted_columns = ', '.join(f'"{c}"' for c in columns)
  sql = f'''
    INSERT INTO "{schema_name}"."{table_name}"
    ({quoted_columns})
    VALUES %s
  '''
  try:
    with transaction.atomic():
      connection = connections['stadtbereichskatalog']
      with connection.cursor() as cursor:
        execute_values(cursor, sql, rows)
    return {'success': True, 'inserted_rows': len(rows)}
  except Exception as e:
    return {'success': False, 'errors': [{'row': 0, 'message': str(e)}]}


def make_error(
  *, row_number, target_column, source_column, value, target_type, error_type='conversion', message
):
  """
  assembles verbose data import error

  :param row_number: row number
  :param target_column: target column
  :param source_column: source column
  :param value: value
  :param target_type: target type
  :param error_type: error type
  :param message: error message
  :return: assembled verbose data import error
  """
  return {
    'row': row_number,
    'target_column': target_column,
    'source_column': source_column,
    'value': value,
    'target_type': target_type,
    'error_type': error_type,
    'message': message,
  }


def normalize_value(value):
  """
  normalizes passed value

  :param value: value to be normalized
  :return: passed value, normalized
  """
  if isinstance(value, float):
    if value.is_integer():
      return int(value)
  if value is None:
    return None
  if isinstance(value, str):
    value = value.strip()
    return value if value != '' else None
  # optional: normalize float that are actually integers
  if isinstance(value, float):
    if value.is_integer():
      return int(value)
  return value


def read_xlsx(file_obj):
  """
  reads passed XLSX file

  :param file_obj: XLSX file
  :return: headers and data rows of passed XLSX file
  """
  wb = load_workbook(file_obj, data_only=True)
  ws = wb.active
  rows = list(ws.iter_rows(values_only=True))
  headers = [str(h).strip() if h is not None else '' for h in rows[0]]
  data_rows = rows[1:]
  result = []
  for row in data_rows:
    row_dict = {}
    for i, value in enumerate(row):
      if i < len(headers):
        row_dict[headers[i]] = normalize_value(value)
    result.append(row_dict)
  return headers, result


def read_tabular_file(file_obj):
  """
  reads passed tabular file

  :param file_obj: tabular file
  :return: headers and data rows of passed tabular file
  """
  file_type = detect_file_type(file_obj)
  file_obj.seek(0)
  if file_type == 'xlsx':
    return read_xlsx(file_obj)
  if file_type == 'csv':
    reader, dialect = get_csv_reader(file_obj)
    rows = [{k: normalize_value(v) for k, v in row.items()} for row in reader]
    headers = clean_headers(reader.fieldnames)
    return headers, rows
  raise ValueError('Unsupported file type')


def update_data(schema_name, table_name, pk, changes):
  """
  updates data of row with passed primary key of passed table within passed database schema
  based on passed changes

  :param schema_name: name of database schema
  :param table_name: name of database table
  :param pk: primary key
  :param changes: changes
  :return: results of updating data of row with passed primary key of passed table
  within passed database schema based on passed changes
  """

  # dynamic set clause
  set_parts, set_values = [], []
  for column, value in changes.items():
    set_parts.append(f'"{column}" = %s')
    set_values.append(value)
  set_clause = ', '.join(set_parts)

  # dynamic where clause
  where_parts, where_values = [], []
  print(pk)
  for column, value in pk.items():
    where_parts.append(f'"{column}" = %s')
    where_values.append(value)
  where_clause = ' AND '.join(where_parts)

  # build query
  query = SQL("""
      UPDATE {}.{}
  """).format(Identifier(schema_name), Identifier(table_name))
  query += SQL(' SET ') + SQL(set_clause)
  query += SQL(' WHERE ') + SQL(where_clause)
  values = set_values + where_values

  # execute query
  try:
    with transaction.atomic():
      connection = connections['stadtbereichskatalog']
      with connection.cursor() as cursor:
        cursor.execute(query, values)
    return {'success': True, 'updated_rows': cursor.rowcount}
  except Exception as e:
    return {'success': False, 'errors': [{'row': 0, 'message': str(e)}]}
